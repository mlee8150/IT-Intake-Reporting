"""The rolling history workbook that backs the trend panels (2 and 4) and the
working-hours-by-team panel (3).

Transition emails only live in the mailbox for as long as retention allows,
which won't cover a 12-month cycle-time trend. So every pipeline run appends
(or overwrites, if re-run for the same week) one row here, and the trend
panels read the last N rows back out rather than re-querying the mailbox.
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models import WeeklySnapshot
from ..vocabulary import REVIEW_TEAMS

SHEET_NAME = "weekly_snapshots"
FIXED_COLUMNS = [
    "week_ending",
    "opened",
    "closed",
    "median_cycle_time_days",
    "open_requests_active",
    "resolved_this_week",
    "aging_90_plus",
    "exec_critical_open",
]


class RollingHistoryWorkbook:
    def __init__(self, path: Path):
        self.path = path

    def _open_or_create(self) -> tuple[Workbook, Worksheet]:
        if self.path.exists():
            wb = load_workbook(self.path)
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            # Write header cells directly rather than via ws.append(): on a
            # brand-new sheet, append()'s internal row counter lands on row
            # 2, not row 1, leaving row 1 blank.
            header = [*FIXED_COLUMNS, *(f"hours_{team}" for team in REVIEW_TEAMS)]
            for col_idx, name in enumerate(header, start=1):
                ws.cell(1, col_idx, name)
        return wb, ws

    def append_week(self, snapshot: WeeklySnapshot) -> None:
        """Add this week's row, or overwrite it in place if the pipeline is re-run."""
        self.path.parent.mkdir(parents=True, exist_ok=True)
        wb, ws = self._open_or_create()
        header = [cell.value for cell in ws[1]]

        row_values = {
            "week_ending": snapshot.week_ending.isoformat(),
            "opened": snapshot.opened,
            "closed": snapshot.closed,
            "median_cycle_time_days": snapshot.median_cycle_time_days,
            "open_requests_active": snapshot.open_requests_active,
            "resolved_this_week": snapshot.resolved_this_week,
            "aging_90_plus": snapshot.aging_90_plus,
            "exec_critical_open": snapshot.exec_critical_open,
        }
        for team in REVIEW_TEAMS:
            row_values[f"hours_{team}"] = snapshot.working_hours_by_team.get(team, 0.0)

        existing_row_idx = None
        week_col_idx = header.index("week_ending") + 1
        for row in ws.iter_rows(min_row=2):
            if row[week_col_idx - 1].value == row_values["week_ending"]:
                existing_row_idx = row[0].row
                break

        target_row = existing_row_idx or (ws.max_row + 1)
        for col_idx, col_name in enumerate(header, start=1):
            ws.cell(target_row, col_idx, row_values.get(col_name))

        wb.save(self.path)

    def read_recent_weeks(self, n: int) -> list[WeeklySnapshot]:
        if not self.path.exists():
            return []
        wb = load_workbook(self.path)
        if SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[SHEET_NAME]
        header = [cell.value for cell in ws[1]]
        idx = {name: i for i, name in enumerate(header)}

        snapshots: list[WeeklySnapshot] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["week_ending"]] is None:
                continue
            working_hours = {
                team: row[idx[f"hours_{team}"]] or 0.0
                for team in REVIEW_TEAMS
                if f"hours_{team}" in idx
            }
            snapshots.append(
                WeeklySnapshot(
                    week_ending=_parse_date(row[idx["week_ending"]]),
                    opened=row[idx["opened"]] or 0,
                    closed=row[idx["closed"]] or 0,
                    median_cycle_time_days=row[idx["median_cycle_time_days"]] or 0.0,
                    working_hours_by_team=working_hours,
                    open_requests_active=row[idx["open_requests_active"]] or 0,
                    resolved_this_week=row[idx["resolved_this_week"]] or 0,
                    aging_90_plus=row[idx["aging_90_plus"]] or 0,
                    exec_critical_open=row[idx["exec_critical_open"]] or 0,
                )
            )
        snapshots.sort(key=lambda s: s.week_ending)
        return snapshots[-n:]


def _parse_date(value: str | date | datetime) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))
