"""Persisted "latest known transition per sub-task," across runs.

Panels 3 and 7 both need, for every currently-open review sub-task, the
transition that put it into its current status — but each run only fetches
transition emails for the current run's ~7-day window (see
pipeline._fetch_transition_events). A sub-task that hasn't moved in longer
than that has no event in this run's fetch at all, so without this store
it's invisible to those panels, not merely undercounted.

This file is our own record, not a re-derivable cache: each run merges
whatever new transitions it just fetched from Outlook into whatever this
file already knows, keeping the newest transition per ticket, and writes
the result back — so a sub-task keeps showing up here run after run even
through weeks with zero activity on it, until it actually moves again (or
is filtered out downstream as terminal — see vocabulary.SUBTASK_TERMINAL_STATUSES).
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models import TransitionEvent

SHEET_NAME = "subtask_latest_transitions"
COLUMNS = [
    "ticket_key",
    "review_team",
    "from_status",
    "to_status",
    "changed_at",
    "changed_by",
    "raw_subject",
]


class SubtaskTransitionStore:
    def __init__(self, path: Path):
        self.path = path

    def _open_or_create(self) -> tuple[Workbook, Worksheet]:
        if self.path.exists():
            wb = load_workbook(self.path)
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            # Write header cells directly rather than via ws.append(): on a
            # brand-new sheet, append()'s internal row counter lands on row
            # 2, not row 1, leaving row 1 blank.
            for col_idx, name in enumerate(COLUMNS, start=1):
                ws.cell(1, col_idx, name)
        return wb, ws

    def read_all(self) -> list[TransitionEvent]:
        if not self.path.exists():
            return []
        wb = load_workbook(self.path)
        if SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[SHEET_NAME]
        header = [cell.value for cell in ws[1]]
        idx = {name: i for i, name in enumerate(header)}

        events: list[TransitionEvent] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["ticket_key"]] is None:
                continue
            events.append(
                TransitionEvent(
                    ticket_key=row[idx["ticket_key"]],
                    is_parent_request=False,
                    review_team=row[idx["review_team"]],
                    from_status=row[idx["from_status"]],
                    to_status=row[idx["to_status"]],
                    # Excel can't store tzinfo (see _to_naive_utc) — every
                    # value written here was normalized to UTC first, so
                    # reattach UTC rather than leave it naive.
                    changed_at=row[idx["changed_at"]].replace(tzinfo=timezone.utc),
                    changed_by=row[idx["changed_by"]],
                    raw_subject=row[idx["raw_subject"]] or "",
                )
            )
        return events

    def merge_and_save(self, new_events: list[TransitionEvent]) -> list[TransitionEvent]:
        """`new_events` should be this run's latest transition per sub-task
        (see pipeline._latest_transition_per_ticket). Merges into whatever's
        already stored — newer `changed_at` per ticket wins — persists the
        result, and returns the full merged list."""
        merged: dict[str, TransitionEvent] = {e.ticket_key: e for e in self.read_all()}
        for event in new_events:
            current = merged.get(event.ticket_key)
            if current is None or event.changed_at > current.changed_at:
                merged[event.ticket_key] = event

        self.path.parent.mkdir(parents=True, exist_ok=True)
        wb, ws = self._open_or_create()
        # Rewrite the sheet from scratch each time — simpler than in-place
        # row updates, and this file is never meant to be hand-edited.
        for row in list(ws.iter_rows(min_row=2)):
            for cell in row:
                cell.value = None
        for row_idx, event in enumerate(merged.values(), start=2):
            ws.cell(row_idx, 1, event.ticket_key)
            ws.cell(row_idx, 2, event.review_team)
            ws.cell(row_idx, 3, event.from_status)
            ws.cell(row_idx, 4, event.to_status)
            ws.cell(row_idx, 5, _to_naive_utc(event.changed_at))
            ws.cell(row_idx, 6, event.changed_by)
            ws.cell(row_idx, 7, event.raw_subject)
        wb.save(self.path)

        return list(merged.values())


def _to_naive_utc(value: datetime) -> datetime:
    # Excel/openpyxl rejects tz-aware datetimes outright — normalize to the
    # equivalent UTC instant, then drop tzinfo, rather than naively
    # `.replace(tzinfo=None)`-ing whatever offset it happened to carry
    # (which would silently shift the wall-clock value).
    if value.tzinfo is not None:
        value = value.astimezone(timezone.utc)
    return value.replace(tzinfo=None)
